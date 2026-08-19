from io import BytesIO
from urllib.parse import quote

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TRANSLIT = str.maketrans(
    "абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ ",
    "abvgdeejzijklmnoprstufxccssiieeuaABVGDEEJZIJKLMNOPRSTUFXCCSSIIEEUA_",
)


def make_wb(sheet_name, headers, col_widths, rows_data, kinds=None, formats=None):
    def get_cell_format(val, col_idx=None):
        if formats and col_idx and col_idx in formats:
            return formats[col_idx]

        is_num = isinstance(val, (int, float)) and not isinstance(val, bool)

        if not is_num:
            return None

        if isinstance(val, int) or (isinstance(val, float) and val.is_integer()):
            return "0"

        return "0.00"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.freeze_panes = "A2"

    hdr_font = Font(name="Arial", size=9, bold=True, color="000000")
    hdr_fill = PatternFill("solid", fgColor="FFFFFF")
    hdr_aln = Alignment(horizontal="center", wrap_text=True)

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = _BORDER
        cell.alignment = hdr_aln
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(rows_data, 2):
        kind = kinds[ri - 2] if kinds else "normal"
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = _BORDER
            is_num = isinstance(val, (int, float)) and not isinstance(val, bool)
            if kind == "total":
                cell.font = Font(name="Arial", size=9, bold=True, color="000000")
                cell.fill = PatternFill("solid", fgColor="C0C0C0")
                cell.alignment = Alignment(horizontal="right" if is_num else "left")
            elif kind == "subtotal":
                cell.font = Font(name="Arial", size=9, bold=True)
                cell.fill = PatternFill("solid", fgColor="C0C0C0")
                cell.alignment = Alignment(horizontal="right" if is_num else "left")
            else:
                cell.font = Font(name="Arial", size=9)
                if is_num:
                    cell.alignment = Alignment(horizontal="right")
            if is_num:
                cell.number_format = get_cell_format(val, ci)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_response(data, filename_ru):
    today = timezone.localdate().strftime("%d_%m_%Y")
    fname_ascii = f"{filename_ru.translate(_TRANSLIT)}_{today}.xlsx"
    fname_utf8 = quote(f"{filename_ru}_{today}.xlsx")
    response = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=\"{fname_ascii}\"; filename*=UTF-8''{fname_utf8}"
    )
    return response
