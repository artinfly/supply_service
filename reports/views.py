from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.db import connection
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.management import call_command
from collections import defaultdict
from io import StringIO
from urllib.parse import quote
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile, os, io, zipfile, re


CONCLUDED  = ('Исполняется', 'Возвращен на уточнение', 'На согласовании', 'Подписан', 'На утверждении')
NOT_CONCL  = ('Черновик',)
TERMINATED = ('Расторгнут',)
YEARS      = [2025, 2026, 2027]
YEAR_COL   = {str(y): f'y{str(y)[2:]}' for y in YEARS}

_THIN     = Side(style='thin')
_BORDER   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TRANSLIT = str.maketrans(
    'абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ ',
    'abvgdeejzijklmnoprstufxccssiieeuaABVGDEEJZIJKLMNOPRSTUFXCCSSIIEEUA_'
)

_HISTORY_JOIN = """
    FROM contracts_history ch
    LEFT JOIN igk_stat_data isd ON ch.hash = digest(
        concat(isd.igk, isd.c_agent, isd.contract, isd.item,
               isd."order", TRIM(isd.stage), isd.plan_date), 'md5')
"""
_HISTORY_GROUP_BASE = """
    isd.igk, isd.c_agent, isd.cfo, isd.contract, isd.item,
    isd.payment_type, isd."order", ch.update_date, ch.upload_date, isd.c_date
"""


def login_view(request):
    if request.user.is_authenticated:
        return redirect('/reports/')
    error = False
    if request.method == 'POST':
        user = authenticate(request, username=request.POST.get('username'), password=request.POST.get('password'))
        if user:
            login(request, user)
            return redirect('/reports/')
        error = True
    return render(request, 'reports/login.html', {'error': error})


def logout_view(request):
    logout(request)
    return redirect('/login/')


def is_operator(user):
    return user.is_superuser or user.groups.filter(name='operator').exists()


def _years_context(request=None):
    ctx = {'years': YEARS, 'year_cols': [(y, f'y{str(y)[2:]}') for y in YEARS]}
    if request:
        ctx['is_operator'] = is_operator(request.user)
    return ctx


def _query_to_json(sql, params=None):
    with connection.cursor() as cur:
        cur.execute(sql, params or [])
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    for row in rows:
        for k, v in row.items():
            if hasattr(v, '__float__'):
                row[k] = float(v)
    return JsonResponse(rows, safe=False, json_dumps_params={'ensure_ascii': False})


def _status_list(statuses):
    return ', '.join(f"'{s}'" for s in statuses)


def _xe(s):
    if s is None:
        return ''
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


def _make_wb(sheet_name, headers, col_widths, rows_data, kinds=None):
    wb = Workbook()
    ws = wb.active
    ws.title        = sheet_name
    ws.freeze_panes = 'A2'

    hdr_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1A3A5C')
    hdr_aln  = Alignment(horizontal='center', wrap_text=True)

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.border    = _BORDER
        cell.alignment = hdr_aln
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(rows_data, 2):
        kind = kinds[ri - 2] if kinds else 'normal'
        for ci, val in enumerate(row, 1):
            cell        = ws.cell(row=ri, column=ci, value=val)
            cell.border = _BORDER
            is_num      = isinstance(val, (int, float)) and not isinstance(val, bool)
            if kind == 'total':
                cell.font      = Font(name='Arial', size=9, bold=True, color='FFFFFF')
                cell.fill      = PatternFill('solid', fgColor='1A3A5C')
                cell.alignment = Alignment(horizontal='right' if is_num else 'left')
            elif kind == 'subtotal':
                cell.font      = Font(name='Arial', size=9, bold=True)
                cell.fill      = PatternFill('solid', fgColor='FDE9D4')
                cell.alignment = Alignment(horizontal='right' if is_num else 'left')
            else:
                cell.font = Font(name='Arial', size=9)
                if is_num:
                    cell.alignment = Alignment(horizontal='right')
            if is_num:
                cell.number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_response(data, filename_ru):
    today       = date.today().strftime('%d_%m_%Y')
    fname_ascii = f'{filename_ru.translate(_TRANSLIT)}_{today}.xlsx'
    fname_utf8  = quote(f'{filename_ru}_{today}.xlsx')
    response    = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename=\"{fname_ascii}\"; filename*=UTF-8''{fname_utf8}"
    return response


# ── Page views ────────────────────────────────────────────────────────────────

@login_required
def index(request):
    return render(request, 'reports/index.html', _years_context(request))


@login_required
def kdr_table(request, year):
    ctx = _years_context(request)
    ctx['year'] = year
    return render(request, 'reports/kdr_table.html', ctx)


@login_required
def igk_concluded_table(request, year):
    ctx = _years_context(request)
    ctx.update({'year': year, 'report_type': 'concluded', 'title': f'ИГК {year} — Заключённые'})
    return render(request, 'reports/igk_table.html', ctx)


@login_required
def igk_not_concluded_table(request, year):
    ctx = _years_context(request)
    ctx.update({'year': year, 'report_type': 'not_concluded', 'title': f'ИГК {year} — Незаключённые'})
    return render(request, 'reports/igk_table.html', ctx)


@login_required
def igk_terminated_table(request, year):
    ctx = _years_context(request)
    ctx.update({'year': year, 'report_type': 'terminated', 'title': f'ИГК {year} — Расторгнутые'})
    return render(request, 'reports/igk_table.html', ctx)


@login_required
def all_contracts_table(request):
    with connection.cursor() as cur:
        cur.execute("SELECT DISTINCT RIGHT(igk, 4) FROM igk_stat_data WHERE igk IS NOT NULL ORDER BY RIGHT(igk, 4)")
        igk_list = [r[0] for r in cur.fetchall()]
    ctx = _years_context(request)
    ctx.update({
        'igk_list':            igk_list,
        'concluded_statuses':  list(CONCLUDED),
        'not_concl_statuses':  list(NOT_CONCL),
        'terminated_statuses': list(TERMINATED),
    })
    return render(request, 'reports/all_contracts.html', ctx)


@login_required
def history_status_table(request):
    return render(request, 'reports/history_status.html', _years_context(request))


@login_required
def history_plan_table(request):
    return render(request, 'reports/history_plan.html', _years_context(request))


@login_required
def history_fact_table(request):
    return render(request, 'reports/history_fact.html', _years_context(request))


@login_required
def contract_dupes_table(request):
    return render(request, 'reports/contract_dupes.html', _years_context(request))


@login_required
def export_page(request):
    with connection.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT c_agent FROM igk_stat_data
            WHERE c_agent IS NOT NULL AND TRIM(c_agent) != ''
            ORDER BY c_agent
        """)
        agents = [r[0] for r in cur.fetchall()]
    ctx = _years_context(request)
    ctx['agents'] = agents
    return render(request, 'reports/export.html', ctx)


@login_required
def upload_excel(request):
    if not is_operator(request.user):
        return JsonResponse({'error': 'forbidden'}, status=403)
    result = None
    if request.method == 'POST' and request.FILES.get('excel_file'):
        f = request.FILES['excel_file']
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            for chunk in f.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        try:
            out = StringIO()
            call_command('import_excel', tmp_path, stdout=out)
            call_command('normalize_staging', stdout=out)
            result = out.getvalue()
            messages.success(request, 'Файл успешно загружен и нормализован')
        except Exception as e:
            messages.error(request, f'Ошибка: {e}')
            result = str(e)
        finally:
            os.unlink(tmp_path)
    ctx = _years_context(request)
    ctx['result'] = result
    return render(request, 'reports/upload.html', ctx)


# ── API views ─────────────────────────────────────────────────────────────────

@login_required
def api_kdr(request, year):
    yc = YEAR_COL.get(year)
    if not yc:
        return JsonResponse({'error': 'invalid year'}, status=400)
    cl  = _status_list(CONCLUDED)
    nl  = _status_list(NOT_CONCL)
    sql = f"""
    SELECT igk,
        COUNT(DISTINCT contract) FILTER (WHERE "order" IS NOT NULL AND "order" != '')                           AS orders,
        ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE "order" IS NOT NULL AND "order" != ''), 0)/1e6 AS numeric), 2) AS order_sum,
        COUNT(DISTINCT contract) FILTER (WHERE "order" != '' AND status IN ({cl}))                              AS count_concluded,
        ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE "order" != '' AND status IN ({cl})), 0)/1e6 AS numeric), 2) AS concluded_order_sum,
        COUNT(DISTINCT contract) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status!='Расторгнут')      AS count_curr_year,
        ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status!='Расторгнут'), 0)/1e6 AS numeric), 2) AS order_sum_curr_year,
        COUNT(DISTINCT contract) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status IN ({cl}))          AS count_concluded_curr_year,
        ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status IN ({cl})), 0)/1e6 AS numeric), 2) AS concluded_order_sum_curr_year,
        ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND payment_type='Аванс' AND status!='Расторгнут'), 0)/1e6 AS numeric), 2) AS pp_sum_plan,
        ROUND(CAST(COALESCE(SUM(fact) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND payment_type='Аванс' AND status!='Расторгнут'), 0)/1e6 AS numeric), 2) AS pp_sum_fact,
        COUNT(DISTINCT contract) FILTER (WHERE "order" != '' AND status IN ({nl}) AND {yc}=TRUE)                AS count_not_concluded_curr_year,
        ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status IN ({nl})), 0)/1e6 AS numeric), 2) AS not_concluded_order_sum_curr_year,
        CASE WHEN COUNT(DISTINCT contract) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status!='Расторгнут') = 0 THEN 0
             ELSE CAST(COUNT(DISTINCT contract) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status IN ({cl})) * 100.0
                  / NULLIF(COUNT(DISTINCT contract) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status!='Расторгнут'), 0) AS int)
        END AS count_concluded_percent_curr_year,
        CASE WHEN COALESCE(SUM(plan) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status!='Расторгнут'), 0) = 0 THEN 0
             ELSE CAST(COALESCE(SUM(plan) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status IN ({cl})), 0) * 100.0
                  / NULLIF(SUM(plan) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND status!='Расторгнут'), 0) AS int)
        END AS order_sum_percent_curr_year,
        CASE WHEN COALESCE(SUM(plan) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND payment_type='Аванс' AND status!='Расторгнут'), 0) = 0 THEN 0
             ELSE CAST(COALESCE(SUM(fact) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND payment_type='Аванс' AND status!='Расторгнут'), 0)
                  / NULLIF(SUM(plan) FILTER (WHERE "order" IS NOT NULL AND {yc}=TRUE AND payment_type='Аванс' AND status!='Расторгнут'), 0) * 100 AS int)
        END AS pp_percent
    FROM igk_stat_data
    GROUP BY igk ORDER BY igk
    """
    return _query_to_json(sql)


def _igk_stat_sql(yc, statuses):
    sl = _status_list(statuses)
    return f"""
    SELECT igk,
        ROUND(CAST(COALESCE(SUM(plan), 0) AS numeric), 2)                                                                              AS spec_sum,
        ROUND(CAST(COALESCE(SUM(plan)  FILTER (WHERE payment_type='Аванс'), 0) AS numeric), 2)                                         AS pp_sum,
        ROUND(CAST(COALESCE(SUM(plan)  FILTER (WHERE payment_type='Аванс'), 0)*100.0 / NULLIF(SUM(plan),0) AS numeric), 0)             AS pp_percent,
        ROUND(CAST(COALESCE(SUM(fact)  FILTER (WHERE payment_type='Аванс'), 0) AS numeric), 2)                                         AS pp_fact,
        ROUND(CAST(COALESCE(SUM(fact)  FILTER (WHERE payment_type='Аванс'), 0)*100.0 / NULLIF(SUM(plan),0) AS numeric), 0)             AS fact_percent,
        ROUND(CAST(COALESCE(SUM(plan)  FILTER (WHERE payment_type='Аванс' AND plan>=0), 0)
                 - COALESCE(SUM(fact)  FILTER (WHERE payment_type='Аванс' AND plan>=0), 0) AS numeric), 2)                             AS pp_remain,
        ROUND(CAST((COALESCE(SUM(plan) FILTER (WHERE payment_type='Аванс'), 0)
                  - COALESCE(SUM(fact) FILTER (WHERE payment_type='Аванс'), 0))*100.0 / NULLIF(SUM(plan),0) AS numeric), 0)            AS remain_percent,
        COUNT(*)                                                                                                                        AS pp_quantity
    FROM igk_stat_data
    WHERE {yc} = TRUE AND status IN ({sl})
    GROUP BY igk ORDER BY igk
    """


def _igk_stat_total_sql(yc, statuses):
    sl = _status_list(statuses)
    return f"""
    SELECT 'ИТОГО' AS igk,
        ROUND(CAST(COALESCE(SUM(plan), 0) AS numeric), 2)                                                                              AS spec_sum,
        ROUND(CAST(COALESCE(SUM(plan)  FILTER (WHERE payment_type='Аванс'), 0) AS numeric), 2)                                         AS pp_sum,
        ROUND(CAST(COALESCE(SUM(plan)  FILTER (WHERE payment_type='Аванс'), 0)*100.0 / NULLIF(SUM(plan),0) AS numeric), 0)             AS pp_percent,
        ROUND(CAST(COALESCE(SUM(fact)  FILTER (WHERE payment_type='Аванс'), 0) AS numeric), 2)                                         AS pp_fact,
        ROUND(CAST(COALESCE(SUM(fact)  FILTER (WHERE payment_type='Аванс'), 0)*100.0 / NULLIF(SUM(plan),0) AS numeric), 0)             AS fact_percent,
        ROUND(CAST(COALESCE(SUM(plan)  FILTER (WHERE payment_type='Аванс' AND plan>=0), 0)
                 - COALESCE(SUM(fact)  FILTER (WHERE payment_type='Аванс' AND plan>=0), 0) AS numeric), 2)                             AS pp_remain,
        ROUND(CAST((COALESCE(SUM(plan) FILTER (WHERE payment_type='Аванс'), 0)
                  - COALESCE(SUM(fact) FILTER (WHERE payment_type='Аванс'), 0))*100.0 / NULLIF(SUM(plan),0) AS numeric), 0)            AS remain_percent,
        COUNT(*)                                                                                                                        AS pp_quantity
    FROM igk_stat_data
    WHERE {yc} = TRUE AND status IN ({sl})
    """


def _igk_stat_response(year, statuses):
    yc = YEAR_COL.get(year)
    if not yc:
        return JsonResponse({'error': 'invalid year'}, status=400)
    with connection.cursor() as cur:
        cur.execute(_igk_stat_sql(yc, statuses))
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        cur.execute(_igk_stat_total_sql(yc, statuses))
        rows.append(dict(zip(cols, cur.fetchone())))
    for row in rows:
        for k, v in row.items():
            if hasattr(v, '__float__'):
                row[k] = float(v)
    return JsonResponse(rows, safe=False, json_dumps_params={'ensure_ascii': False})


@login_required
def api_igk_concluded(request, year):
    return _igk_stat_response(year, CONCLUDED)


@login_required
def api_igk_not_concluded(request, year):
    return _igk_stat_response(year, NOT_CONCL)


@login_required
def api_igk_terminated(request, year):
    return _igk_stat_response(year, TERMINATED)


@login_required
def api_all_contracts(request):
    agent       = request.GET.get('agent', '').strip()
    igk_filter  = request.GET.get('igk', '').strip()
    statuses    = request.GET.getlist('status')
    year_filter = request.GET.get('year', '').strip()

    conditions = ["payment_type IS NOT NULL AND TRIM(payment_type) != ''"]
    params = []
    if agent:
        conditions.append("c_agent ILIKE %s")
        params.append(f'%{agent}%')
    if igk_filter:
        conditions.append("igk LIKE %s")
        params.append(f'%{igk_filter}')
    if statuses:
        conditions.append(f"status IN ({','.join(['%s']*len(statuses))})")
        params.extend(statuses)
    if year_filter in YEAR_COL:
        conditions.append(f"{YEAR_COL[year_filter]} = TRUE")

    where      = 'WHERE ' + ' AND '.join(conditions)
    sql_detail = f"""
        SELECT igk, c_agent, contract, status,
            COALESCE(payment_type,'ИНОЕ') AS payment_type,
            item, "order", TRIM(stage) AS stage, y25, y26, y27,
            ROUND(CAST(SUM(plan) AS numeric), 2) AS spec_sum,
            ROUND(CAST(SUM(plan)  FILTER (WHERE payment_type='Аванс') AS numeric), 2) AS pp_sum,
            ROUND(CAST(SUM(fact)  FILTER (WHERE payment_type='Аванс') AS numeric), 2) AS pp_fact,
            ROUND(CAST(SUM(plan) - SUM(COALESCE(fact,0)) AS numeric), 2) AS pp_remain,
            0 AS is_subtotal
        FROM igk_stat_data {where}
        GROUP BY igk, c_agent, contract, status, payment_type, item, "order", stage, y25, y26, y27
        ORDER BY igk NULLS LAST, contract, payment_type
    """
    sql_total  = f"""
        SELECT igk, c_agent, contract, status,
            'ИТОГО' AS payment_type,
            item, "order", NULL AS stage, y25, y26, y27,
            ROUND(CAST(SUM(plan) AS numeric), 2) AS spec_sum,
            ROUND(CAST(SUM(plan)  FILTER (WHERE payment_type='Аванс') AS numeric), 2) AS pp_sum,
            ROUND(CAST(SUM(fact)  FILTER (WHERE payment_type='Аванс') AS numeric), 2) AS pp_fact,
            ROUND(CAST(SUM(plan) - SUM(COALESCE(fact,0)) AS numeric), 2) AS pp_remain,
            1 AS is_subtotal
        FROM igk_stat_data {where}
        GROUP BY igk, c_agent, contract, status, item, "order", y25, y26, y27
        ORDER BY igk NULLS LAST, contract
    """
    with connection.cursor() as cur:
        cur.execute(sql_detail, params)
        cols   = [c[0] for c in cur.description]
        detail = [dict(zip(cols, r)) for r in cur.fetchall()]
        cur.execute(sql_total, params)
        totals = {(r[0], r[2], r[6]): dict(zip(cols, r)) for r in cur.fetchall()}

    groups = defaultdict(list)
    for row in detail:
        groups[(row['igk'], row['contract'], row['order'])].append(row)

    result = []
    for key, rows in groups.items():
        result.extend(rows)
        if key in totals:
            result.append(totals[key])

    for row in result:
        for k, v in row.items():
            if hasattr(v, '__float__'):
                row[k] = float(v)
    return JsonResponse(result, safe=False, json_dumps_params={'ensure_ascii': False})


@login_required
def api_history_status(request):
    sql = f"""
        SELECT RIGHT(isd.igk, 4) AS igk, isd.c_agent, isd.cfo, isd.contract,
            ch.old_status, ch.new_status, isd.payment_type, isd.item,
            ROUND(CAST(SUM(isd.plan) AS numeric), 2) AS plan_sum,
            ROUND(CAST(SUM(isd.fact) AS numeric), 2) AS fact_sum,
            ch.update_date, ch.upload_date, isd.c_date
        {_HISTORY_JOIN}
        WHERE ch.old_status IS NOT NULL
        GROUP BY {_HISTORY_GROUP_BASE}, ch.old_status, ch.new_status
        ORDER BY ch.update_date DESC NULLS LAST
    """
    return _query_to_json(sql)


@login_required
def api_history_plan(request):
    sql = f"""
        SELECT RIGHT(isd.igk, 4) AS igk, isd.c_agent, isd.cfo, isd.contract,
            isd.payment_type, isd.item, ch.old_plan, ch.new_plan,
            ch.plan_changed_date, isd.c_date
        {_HISTORY_JOIN}
        WHERE ch.plan_changed_date IS NOT NULL
        GROUP BY {_HISTORY_GROUP_BASE}, ch.old_plan, ch.new_plan, ch.plan_changed_date
        ORDER BY ch.plan_changed_date DESC NULLS LAST
    """
    return _query_to_json(sql)


@login_required
def api_history_fact(request):
    sql = f"""
        SELECT RIGHT(isd.igk, 4) AS igk, isd.c_agent, isd.cfo, isd.contract,
            isd.payment_type, isd.item, ch.old_fact, ch.new_fact,
            ch.fact_changed_date, isd.c_date
        {_HISTORY_JOIN}
        WHERE ch.fact_changed_date IS NOT NULL
        GROUP BY {_HISTORY_GROUP_BASE}, ch.old_fact, ch.new_fact, ch.fact_changed_date
        ORDER BY ch.fact_changed_date DESC NULLS LAST
    """
    return _query_to_json(sql)


@login_required
def api_contract_dupes(request):
    sql = """
        SELECT c_agent, contract, item, "order", TRIM(stage) AS stage,
               plan_date, encode(digest(concat(
               c_agent, contract, item, "order", TRIM(stage), plan_date), 'md5'), 'hex') AS hash
        FROM igk_stat_data
        GROUP BY igk, c_agent, contract, item, "order", stage, plan_date
        HAVING COUNT(*) > 1
        ORDER BY contract, c_agent
    """
    return _query_to_json(sql)


@login_required
def api_igk_detail(request, year, igk):
    yc = YEAR_COL.get(year)
    if not yc:
        return JsonResponse({'error': 'invalid year'}, status=400)
    report_type = request.GET.get('type', 'concluded')
    statuses    = {'concluded': CONCLUDED, 'not_concluded': NOT_CONCL}.get(report_type, TERMINATED)
    sl          = _status_list(statuses)
    sql = f"""
        SELECT contract, c_agent, status,
            COALESCE(payment_type,'ИНОЕ') AS payment_type,
            item, "order", TRIM(stage) AS stage,
            ROUND(CAST(SUM(plan) AS numeric), 2) AS plan_sum,
            ROUND(CAST(SUM(fact) AS numeric), 2) AS fact_sum,
            ROUND(CAST(SUM(plan)-SUM(COALESCE(fact,0)) AS numeric), 2) AS remain
        FROM igk_stat_data
        WHERE igk LIKE %s AND {yc}=TRUE AND status IN ({sl})
          AND payment_type IS NOT NULL AND TRIM(payment_type) != ''
        GROUP BY contract, c_agent, status, payment_type, item, "order", stage
        ORDER BY contract, payment_type
    """
    return _query_to_json(sql, [f'%{igk}'])


@login_required
def api_contracts_by_agent(request):
    year  = request.GET.get('year', '').strip()
    agent = request.GET.get('agent', '').strip()
    yc    = YEAR_COL.get(year)
    if not yc or not agent:
        return JsonResponse([], safe=False)
    sql = f"""
        SELECT igk, c_agent, contract, status, payment_type,
            item, "order", TRIM(stage) AS stage,
            ROUND(CAST(SUM(plan) AS numeric), 2) AS plan,
            ROUND(CAST(SUM(COALESCE(fact,0)) AS numeric), 2) AS fact,
            ROUND(CAST(SUM(plan)-SUM(COALESCE(fact,0)) AS numeric), 2) AS remain
        FROM igk_stat_data
        WHERE {yc}=TRUE AND is_deleted=FALSE
          AND c_agent ILIKE %s
          AND payment_type IS NOT NULL AND TRIM(payment_type) != ''
          AND contract IS NOT NULL AND TRIM(contract) != ''
        GROUP BY igk, c_agent, contract, status, payment_type, item, "order", stage
        ORDER BY igk, contract, payment_type
    """
    return _query_to_json(sql, [f'%{agent}%'])


# ── Export ────────────────────────────────────────────────────────────────────

def _export_simple(sql, params, name, headers, col_widths):
    with connection.cursor() as cur:
        cur.execute(sql, params)
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    return _xlsx_response(_make_wb(name, headers, col_widths, [[row[c] for c in cols] for row in rows]), name)


@login_required
def export_history_status(request):
    sql = f"""
        SELECT RIGHT(isd.igk, 4) AS igk, isd.c_agent, isd.cfo, isd.contract,
            ch.old_status, ch.new_status, isd.payment_type, isd.item,
            ROUND(CAST(SUM(isd.plan) AS numeric), 2) AS plan_sum,
            ROUND(CAST(SUM(isd.fact) AS numeric), 2) AS fact_sum,
            ch.update_date, ch.upload_date, isd.c_date
        {_HISTORY_JOIN}
        WHERE ch.old_status IS NOT NULL
        GROUP BY {_HISTORY_GROUP_BASE}, ch.old_status, ch.new_status
        ORDER BY ch.update_date DESC NULLS LAST
    """
    return _export_simple(sql, [], 'история_статусов',
        ['ИГК', 'Контрагент', 'ЦФО', 'Договор', 'Статус (было)', 'Статус (стало)',
         'Тип платежа', 'Предмет', 'План, руб.', 'Факт, руб.',
         'Дата изменения', 'Дата загрузки', 'Дата договора'],
        [10, 40, 8, 50, 25, 25, 15, 50, 16, 16, 12, 12, 12])


@login_required
def export_history_plan(request):
    sql = f"""
        SELECT RIGHT(isd.igk, 4) AS igk, isd.c_agent, isd.cfo, isd.contract,
            isd.payment_type, isd.item, ch.old_plan, ch.new_plan,
            ch.plan_changed_date, isd.c_date
        {_HISTORY_JOIN}
        WHERE ch.plan_changed_date IS NOT NULL
        GROUP BY {_HISTORY_GROUP_BASE}, ch.old_plan, ch.new_plan, ch.plan_changed_date
        ORDER BY ch.plan_changed_date DESC NULLS LAST
    """
    return _export_simple(sql, [], 'история_плана',
        ['ИГК', 'Контрагент', 'ЦФО', 'Договор', 'Тип платежа', 'Предмет',
         'План (было), руб.', 'План (стало), руб.', 'Дата изменения', 'Дата договора'],
        [10, 40, 8, 50, 15, 50, 16, 16, 12, 12])


@login_required
def export_history_fact(request):
    sql = f"""
        SELECT RIGHT(isd.igk, 4) AS igk, isd.c_agent, isd.cfo, isd.contract,
            isd.payment_type, isd.item, ch.old_fact, ch.new_fact,
            ch.fact_changed_date, isd.c_date
        {_HISTORY_JOIN}
        WHERE ch.fact_changed_date IS NOT NULL
        GROUP BY {_HISTORY_GROUP_BASE}, ch.old_fact, ch.new_fact, ch.fact_changed_date
        ORDER BY ch.fact_changed_date DESC NULLS LAST
    """
    return _export_simple(sql, [], 'история_факта',
        ['ИГК', 'Контрагент', 'ЦФО', 'Договор', 'Тип платежа', 'Предмет',
         'Факт (было), руб.', 'Факт (стало), руб.', 'Дата изменения', 'Дата договора'],
        [10, 40, 8, 50, 15, 50, 16, 16, 12, 12])


@login_required
def export_contract_dupes(request):
    sql = """
        SELECT c_agent, contract, item, "order", TRIM(stage) AS stage,
               plan_date, encode(digest(concat(
               c_agent, contract, item, "order", TRIM(stage), plan_date), 'md5'), 'hex') AS hash
        FROM igk_stat_data
        GROUP BY igk, c_agent, contract, item, "order", stage, plan_date
        HAVING COUNT(*) > 1
        ORDER BY contract, c_agent
    """
    return _export_simple(sql, [], 'дубли_договоров',
        ['Контрагент', 'Договор', 'Предмет', 'Заказ', 'Этап', 'Дата плана', 'Хеш'],
        [40, 50, 50, 20, 15, 12, 12])


@login_required
def export_kdr(request, year):
    yc = YEAR_COL.get(str(year))
    if not yc:
        return JsonResponse({'error': 'invalid year'}, status=400)

    cl = _status_list(CONCLUDED)
    nl = _status_list(NOT_CONCL)
    yy = str(year)[2:]

    sql = f"""
        SELECT MAX(igk) AS igk, cfo,
            COUNT(DISTINCT contract)                                                                                          AS total_count,
            ROUND(CAST(COALESCE(SUM(plan),0)/1e6 AS numeric),2)                                                              AS total_sum,
            COUNT(DISTINCT contract) FILTER (WHERE status IN ({cl}))                                                         AS concl_count,
            ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE status IN ({cl})),0)/1e6 AS numeric),2)                              AS concl_sum,
            COUNT(DISTINCT contract) FILTER (WHERE {yc}=TRUE AND status!='Расторгнут')                                       AS year_count,
            ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE {yc}=TRUE AND status!='Расторгнут'),0)/1e6 AS numeric),2)            AS year_sum,
            COUNT(DISTINCT contract) FILTER (WHERE {yc}=TRUE AND status IN ({cl}))                                           AS year_concl_count,
            ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE {yc}=TRUE AND status IN ({cl})),0)/1e6 AS numeric),2)                AS year_concl_sum,
            COUNT(DISTINCT contract) FILTER (WHERE {yc}=TRUE AND status IN ({nl}))                                           AS year_not_concl_count,
            ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE {yc}=TRUE AND status IN ({nl})),0)/1e6 AS numeric),2)                AS year_not_concl_sum,
            ROUND(CAST(COALESCE(SUM(plan) FILTER (WHERE {yc}=TRUE AND payment_type='Аванс' AND status!='Расторгнут'),0)/1e6 AS numeric),2) AS pp_plan,
            ROUND(CAST(COALESCE(SUM(fact) FILTER (WHERE {yc}=TRUE AND payment_type='Аванс' AND status!='Расторгнут'),0)/1e6 AS numeric),2) AS pp_fact
        FROM igk_stat_data
        WHERE igk IS NOT NULL AND TRIM(igk)!='' AND cfo IS NOT NULL AND TRIM(cfo)!=''
        GROUP BY RIGHT(igk,4), cfo
        ORDER BY RIGHT(igk,4), cfo
    """
    with connection.cursor() as cur:
        cur.execute(sql)
        db_cols     = [c[0] for c in cur.description]
        detail_rows = [dict(zip(db_cols, r)) for r in cur.fetchall()]

    def fv(v):     return float(v or 0)
    def pct(a, b): return round(fv(a) / fv(b) * 100, 1) if fv(b) else 0.0

    def row_vals(r, igk_label, cfo_label):
        yn, ys = fv(r['year_count']), fv(r['year_sum'])
        return [
            igk_label, cfo_label,
            fv(r['total_count']),          fv(r['total_sum']),
            fv(r['concl_count']),          fv(r['concl_sum']),
            fv(r['year_count']),           fv(r['year_sum']),
            fv(r['year_concl_count']),     pct(r['year_concl_count'], yn),
            fv(r['year_concl_sum']),       pct(r['year_concl_sum'], ys),
            fv(r['year_not_concl_count']), fv(r['year_not_concl_sum']),
            fv(r['pp_plan']),              fv(r['pp_fact']),
            pct(r['pp_fact'], r['pp_plan']),
        ]

    def sum_group(rows):
        keys = ['total_count','total_sum','concl_count','concl_sum','year_count','year_sum',
                'year_concl_count','year_concl_sum','year_not_concl_count','year_not_concl_sum','pp_plan','pp_fact']
        return {k: sum(fv(r[k]) for r in rows) for k in keys}

    igk_groups = defaultdict(list)
    for r in detail_rows:
        igk_groups[r['igk']].append(r)

    rows, kinds = [], []
    for igk, grp in igk_groups.items():
        rows.append(row_vals(sum_group(grp), igk, 'Итого')); kinds.append('subtotal')
        for r in grp:
            rows.append(row_vals(r, '', r['cfo'])); kinds.append('normal')
    rows.append(row_vals(sum_group(detail_rows), 'ИТОГО', '')); kinds.append('total')

    HEADERS = [
        'ИГК', 'ЦФО',
        'Всего дог., шт.',         'Сумма всего, млн',
        'Заключено, шт.',          'Сумма заключённых, млн',
        f'Дог. {yy}г., шт.',       f'Сумма {yy}г., млн',
        f'Заключено {yy}г., шт.',  f'% конт. {yy}г. (шт.)',
        f'Сумма зак. {yy}г., млн', f'% конт. {yy}г. (сумма)',
        f'Не зак. {yy}г., шт.',    f'Сумма незак. {yy}г., млн',
        f'АП план {yy}г., млн',    f'АП факт {yy}г., млн',
        f'% АП {yy}г.',
    ]
    COL_W = [10, 6, 12, 14, 12, 16, 12, 14, 12, 12, 16, 14, 12, 16, 14, 14, 12]
    return _xlsx_response(_make_wb(f'КДР {year}', HEADERS, COL_W, rows, kinds), f'кдр_{year}')


@login_required
def export_advances(request, year):
    yc = YEAR_COL.get(str(year))
    if not yc:
        return JsonResponse({'error': 'invalid year'}, status=400)

    sql = f"""
        SELECT MAX(igk) AS igk, MAX(c_agent) AS c_agent, MAX(cfo) AS cfo, contract,
            CASE WHEN MAX(status) IN ('Черновик','Приостановлен') THEN 'Не заключён' ELSE 'Заключён' END AS state,
            MAX(payment_type) AS payment_type, MAX(item) AS item, "order" AS qty,
            ROUND(CAST(SUM(CASE WHEN tolerance>0 THEN plan*(1+tolerance/100.0) ELSE plan END) AS numeric),2) AS spec_sum,
            ROUND(CAST(SUM(CASE WHEN payment_type='Аванс' AND tolerance>0 THEN plan*(1+tolerance/100.0)
                              WHEN payment_type='Аванс' THEN plan ELSE 0 END) AS numeric),2) AS advance_plan,
            ROUND(CAST(SUM(CASE WHEN payment_type='Аванс' THEN COALESCE(fact,0) ELSE 0 END) AS numeric),2) AS advance_fact
        FROM igk_stat_data
        WHERE {yc}=TRUE AND is_deleted=FALSE AND status!='Расторгнут'
          AND payment_type IN ('Аванс','Постоплата')
          AND igk IS NOT NULL AND TRIM(igk)!=''
          AND cfo IS NOT NULL AND TRIM(cfo)!=''
          AND contract IS NOT NULL AND TRIM(contract)!=''
        GROUP BY contract, "order"
        ORDER BY MAX(igk), MAX(cfo), contract, "order"
    """
    with connection.cursor() as cur:
        cur.execute(sql)
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    def fv(v): return float(v) if v is not None else 0.0

    template_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'reports', 'static', 'reports', 'templates', 'templateIGK.xlsx')

    with zipfile.ZipFile(template_path) as z_src:
        orig_s2  = z_src.read('xl/worksheets/sheet2.xml').decode('utf-8')
        orig_ss  = z_src.read('xl/sharedStrings.xml').decode('utf-8')

    cols_xml   = re.search(r'<cols>.*?</cols>', orig_s2, re.DOTALL).group()
    tmpl_sis   = re.findall(r'<si>.*?</si>', orig_ss, re.DOTALL)
    tmpl_strs  = re.findall(r'<t[^>]*>(.*?)</t>', orig_ss)
    tmpl_count = len(tmpl_strs)
    total_rows = len(rows) + 1

    COL_LETTERS = ['A','B','C','D','E','F','G','H','I','J','K']
    HDR_STYLES  = ['27','1','1','1','1','1','1','1','1','4','4']
    DAT_STYLES  = ['28','18','18','18','18','18','18','19','20','21','22']
    TXT_FIELDS  = ['igk','c_agent','cfo','contract','state','payment_type','item','qty']

    ss_list, ss_map = [], {}

    def ss_idx(val):
        s = '' if val is None else str(val)
        if s not in ss_map:
            ss_map[s] = tmpl_count + len(ss_list)
            ss_list.append(s)
        return ss_map[s]

    spec_hdr     = f'Сумма по спецификации на {year} год'
    spec_hdr_idx = tmpl_strs.index(spec_hdr) if spec_hdr in tmpl_strs else ss_idx(spec_hdr)
    hdr_indices  = [0, 1, 2, 3, 4, 5, 6, 9, spec_hdr_idx, 10, 11]

    data_vals = [
        ([ss_idx(row[f]) for f in TXT_FIELDS], fv(row['spec_sum']), fv(row['advance_plan']), fv(row['advance_fact']))
        for row in rows
    ]

    new_ss = '\n'.join([
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        f' count="{tmpl_count + 8*len(rows)}" uniqueCount="{tmpl_count + len(ss_list)}">',
        *tmpl_sis,
        *[f'<si><t{"  xml:space=\"preserve\"" if sv.startswith(" ") or sv.endswith(" ") else ""}>{_xe(sv)}</t></si>' for sv in ss_list],
        '</sst>',
    ]).encode('utf-8')

    xml_lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
        ' xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"'
        ' mc:Ignorable="x14ac" xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac">',
        '<sheetPr><outlinePr summaryRight="0"/><pageSetUpPr autoPageBreaks="0" fitToPage="1"/></sheetPr>',
        f'<dimension ref="A1:K{total_rows}"/>',
        '<sheetViews><sheetView tabSelected="1" workbookViewId="0">'
        '<selection activeCell="A1" sqref="A1"/></sheetView></sheetViews>',
        '<sheetFormatPr defaultColWidth="10.5" defaultRowHeight="11.45" customHeight="1" x14ac:dyDescent="0.2"/>',
        cols_xml, '<sheetData>',
        '<row r="1" spans="1:11" ht="26.1" customHeight="1" x14ac:dyDescent="0.2">',
        *[f'<c r="{col}1" s="{s}" t="s"><v>{idx}</v></c>' for idx, s, col in zip(hdr_indices, HDR_STYLES, COL_LETTERS)],
        '</row>',
        *[line for ri, (txt_idx, spec_sum, adv_plan, adv_fact) in enumerate(data_vals, 2)
          for line in [
              f'<row r="{ri}" spans="1:11" ht="12" customHeight="1" x14ac:dyDescent="0.2">',
              *[f'<c r="{col}{ri}" s="{s}" t="s"><v>{idx}</v></c>' for idx, s, col in zip(txt_idx, DAT_STYLES[:8], COL_LETTERS[:8])],
              f'<c r="I{ri}" s="{DAT_STYLES[8]}"><v>{spec_sum}</v></c>',
              f'<c r="J{ri}" s="{DAT_STYLES[9]}"><v>{adv_plan}</v></c>',
              f'<c r="K{ri}" s="{DAT_STYLES[10]}"><v>{adv_fact}</v></c>',
              '</row>',
          ]],
        '</sheetData>',
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>',
        '</worksheet>',
    ]
    new_s2 = '\n'.join(xml_lines).encode('utf-8')

    buf = io.BytesIO()
    with zipfile.ZipFile(template_path) as z_src:
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z_out:
            for item in z_src.infolist():
                if item.filename == 'xl/worksheets/sheet2.xml':
                    z_out.writestr(item, new_s2)
                elif item.filename == 'xl/sharedStrings.xml':
                    z_out.writestr(item, new_ss)
                elif item.filename == 'xl/worksheets/sheet1.xml':
                    s1 = z_src.read(item.filename).decode('utf-8')
                    s1 = re.sub(r'<row r="(?:[1-9]|1[0-9]|2[01])"[^>]*hidden="1"[^>]*>.*?</row>', '', s1, flags=re.DOTALL)
                    z_out.writestr(item, s1.encode('utf-8'))
                elif item.filename == 'xl/styles.xml':
                    sxml  = z_src.read(item.filename).decode('utf-8')
                    inner = re.search(r'<cellXfs[^>]*>(.*?)</cellXfs>', sxml, re.DOTALL).group(1).strip()
                    parts = [p for p in re.split(r'(?=<xf )', inner) if p.strip().startswith('<xf')]
                    if len(parts) > 28:
                        old28 = parts[28]
                        new28 = re.sub(r'fillId="\d+"', 'fillId="2"', old28)
                        if 'applyFill' not in new28:
                            new28 = new28.replace('<xf ', '<xf applyFill="1" ', 1)
                        sxml = sxml.replace(old28, new28, 1)
                    z_out.writestr(item, sxml.encode('utf-8'))
                elif item.filename == 'xl/pivotTables/pivotTable1.xml':
                    pt_xml = z_src.read(item.filename).decode('utf-8')
                    if 'refreshOnLoad' not in pt_xml:
                        pt_xml = re.sub(r'(<pivotTableDefinition[^>]*)', r'\1 refreshOnLoad="1"', pt_xml, count=1)
                    else:
                        pt_xml = re.sub(r'refreshOnLoad="[01]"', 'refreshOnLoad="1"', pt_xml)
                    pt_xml = re.sub(r'ref="A22:', 'ref="A1:', pt_xml)
                    z_out.writestr(item, pt_xml.encode('utf-8'))
                else:
                    z_out.writestr(item, z_src.read(item.filename))

    return _xlsx_response(buf.getvalue(), f'игк_{year}')


@login_required
def export_contracts_by_agent(request, year):
    yc = YEAR_COL.get(str(year))
    if not yc:
        return JsonResponse({'error': 'invalid year'}, status=400)

    agent      = request.GET.get('agent', '').strip()
    conditions = [
        f"{yc}=TRUE", "is_deleted=FALSE",
        "contract IS NOT NULL AND TRIM(contract)!=''",
        "payment_type IS NOT NULL AND TRIM(payment_type)!=''",
    ]
    params = []
    if agent:
        conditions.append("c_agent ILIKE %s")
        params.append(f'%{agent}%')

    sql = f"""
        SELECT igk, c_agent, cfo, contract, status, payment_type, item,
               "order", TRIM(stage) AS stage,
               ROUND(CAST(SUM(plan) AS numeric),2) AS plan,
               ROUND(CAST(SUM(COALESCE(fact,0)) AS numeric),2) AS fact,
               ROUND(CAST(SUM(plan)-SUM(COALESCE(fact,0)) AS numeric),2) AS remain
        FROM igk_stat_data
        WHERE {' AND '.join(conditions)}
        GROUP BY igk, c_agent, cfo, contract, status, payment_type, item, "order", stage
        ORDER BY igk, c_agent, contract, "order", payment_type
    """
    with connection.cursor() as cur:
        cur.execute(sql, params)
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    HEADERS = ['ИГК','Контрагент','ЦФО','Договор','Состояние','Тип платежа',
               'Предмет','Заказ','Этап', f'План {year}, руб.', f'Факт {year}, руб.','Остаток, руб.']
    COL_W   = [15, 40, 8, 50, 20, 15, 50, 20, 15, 18, 18, 18]
    TXT_FLD = ['igk','c_agent','cfo','contract','status','payment_type','item','order','stage']

    data_rows = [
        [row[f] for f in TXT_FLD] + [float(row['plan'] or 0), float(row['fact'] or 0), float(row['remain'] or 0)]
        for row in rows
    ]
    agent_safe = agent[:30].replace(' ', '_') if agent else ''
    return _xlsx_response(
        _make_wb(f'Договоры {year}', HEADERS, COL_W, data_rows),
        f'контрагент{"_" + agent_safe if agent_safe else ""}_{year}'
    )